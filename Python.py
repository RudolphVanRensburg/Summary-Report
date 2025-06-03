import os
from datetime import datetime
import pandas as pd # type: ignore
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, Alignment, PatternFill # type: ignore

# Today's date for filename formatting (YYYYMMDD)
today = datetime.now().strftime('%Y%m%d')

# Define file paths for each client and document type
clients_config = {
    'Client 1': {
        'Death': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 1\Death.xlsx',
        'Disability': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 1\Disability.xlsx',
        'Underwriting': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 1\Medical Underwriting.xlsx',
    },
    'Client 2': {
        'Death': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 2\Death.xlsx',  # Should be Client 2 path
        'Disability': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 2\Disability.xlsx',  # Should be Client 2 path
        'Underwriting': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 2\Medical Underwriting.xlsx',  # Should be Client 2 path
    },
    'Client 3': {
        'Death': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 3\Death.xlsx',  # Should be Client 3 path
        'Disability': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 3\Disability.xlsx',  # Should be Client 3 path
        'Underwriting': r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\Client 3\Medical Underwriting.xlsx',  # Should be Client 3 path
    }
}

# Output file directory and name
file_directory = r'C:\Users\Rudolph Van Rensburg\Desktop\KHUSA\Clarissa Summary Report\01 - Summary Reports'
filename = f'{today} - Claims & Underwriting Summarized Reporting.xlsx'
file_path = os.path.join(file_directory, filename)

def clean_columns(df):
    "Strip leading/trailing spaces from all column headers."
    df.columns = df.columns.str.strip()
    return df

def process_client_data(paths):
    "Load and analyse data from file paths for a client."
    # Load Excel files into DataFrames
    death_df = clean_columns(pd.read_excel(paths['Death']))
    disability_df = clean_columns(pd.read_excel(paths['Disability']))
    underwriting_df = clean_columns(pd.read_excel(paths['Underwriting']))

    # Claims summary
    funeral_claims = death_df['Claim type'].astype(str).str.contains('funeral', case=False, na=False).sum()
    gla_claims = death_df['Claim type'].astype(str).str.contains('GLA', case=False, na=False).sum()
    finalised_funeral = death_df[death_df['Status of the Claim'].astype(str).str.contains('Paid', case=False, na=False)]['Claim type'].astype(str).str.contains('funeral', case=False, na=False).sum()
    finalised_gla = death_df[death_df['Status of the Claim'].astype(str).str.contains('Paid', case=False, na=False)]['Claim type'].astype(str).str.contains('GLA', case=False, na=False).sum()
    disability_claims = disability_df['Member'].count()

    # Underwriting summary
    gla_requested = underwriting_df[underwriting_df['Benefit'].astype(str).str.contains('GLA', case=False, na=False) & underwriting_df['Decision'].isna()].shape[0]
    disability_requested = underwriting_df[~underwriting_df['Benefit'].astype(str).str.contains('GLA', case=False, na=False) & underwriting_df['Decision'].isna()].shape[0]
    gla_decisioned = underwriting_df[underwriting_df['Benefit'].astype(str).str.contains('GLA', case=False, na=False) & underwriting_df['Decision'].notna()]['Decision'].count()
    disability_decisioned = underwriting_df[~underwriting_df['Benefit'].astype(str).str.contains('GLA', case=False, na=False) & underwriting_df['Decision'].notna()]['Decision'].count()

    return {
        'funeral_claims': funeral_claims,
        'gla_claims': gla_claims,
        'finalised_funeral': finalised_funeral,
        'finalised_gla': finalised_gla,
        'disability_claims': disability_claims,
        'gla_requested': gla_requested,
        'disability_requested': disability_requested,
        'gla_decisioned': gla_decisioned,
        'disability_decisioned': disability_decisioned,
    }

def format_header_cell(cell):
    """Apply consistent formatting to header cells."""
    cell.font = Font(size=14, bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")

def get_sheet_data(stats):
    "Prepare data layout for the worksheet using collected stats."
    return [
        ['Claims', '', '', '', '', '', 'Underwriting', '', '', ''],
        ['', 'Funeral', 'Death', 'Disability', 'Notes', '', '', 'GLA', 'Disability', 'Notes'],
        ['Received', stats['funeral_claims'], stats['gla_claims'], stats['disability_claims'], '', '', 'Requested', stats['gla_requested'], stats['disability_requested'], ''],
        ['Finalised', stats['finalised_funeral'], stats['finalised_gla'], 'N/A', '', '', 'Decisioned', stats['gla_decisioned'], stats['disability_decisioned'], ''],
        [''],
        ['ADDITIONAL FEEDBACK/NOTES']
    ]

# Create new Excel workbook
workbook = Workbook()
workbook.remove(workbook.active)  # Remove default sheet

# Loop through each client to create individual sheets
for client_name, file_paths in clients_config.items():
    # Process data for current client
    stats = process_client_data(file_paths)

    # Add new worksheet
    sheet = workbook.create_sheet(title=client_name)

    # Insert data into the sheet
    for row in get_sheet_data(stats):
        sheet.append(row)

    # Merge header cells and format them
    sheet.merge_cells('A1:E1')
    sheet.merge_cells('G1:J1')
    sheet.merge_cells('A6:J6')
    format_header_cell(sheet['A1'])
    format_header_cell(sheet['G1'])
    format_header_cell(sheet['A6'])

# Save the workbook to file
workbook.save(file_path)
print("File saved:", file_path)